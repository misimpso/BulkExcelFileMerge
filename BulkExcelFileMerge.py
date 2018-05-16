import os
import os.path
from xlrd import open_workbook, XLRDError
from openpyxl import Workbook
import types
from unidecode import unidecode
import getopt, sys
import Tkinter, tkFileDialog
import shutil

def usage():
	print(  'Usage -- \n\n' +
	
			'python BulkExcelFileMerge.py [ --option, --option, ... ]\n' +
			'    > long options\n' +
			'        --col_start - - - - start coloum constraint (default = 1)\n' +
			'        --col_end - - - - - end column constraint (default = None)\n' +
			'        --help  - - - - - - print usage info\n' +
			'        --include_headers - include column titles (row 1) from first file to merged file (default = True, takes = T, F, True, False)\n'
			'        --name  - - - - - - name of merged file (default = merged.xlsx)\n' +
			'        --row_start - - - - start row constaint (default = 4)\n' +
			'        --row_end - - - - - end row constraint (default = None)\n' +
			'    > example :\n' +
			'        > python BulkExcelFileMerge.py --name example.xlsx --include_headers F --row_start 9 --col_end 4'
			'--'
			)

def get_filenames(dirname):
	# Loop through XLS files
	# If file name, before '_', not in dict, add it sans route
	# If file name in dict, check file route against value route, if lower, reassign value to new file
	# Listify dict to get filenames
	
	filenames = []
	filenames_dict = {}

	for file in os.listdir(dirname):
		if os.path.isfile(file) and file[file.index('.'):] == '.xls' and file[:file.index('.')] != 'test' and '~' not in file:
			if file[:file.index('_')] not in filenames_dict:
				filenames_dict[file[:file.index('_')]] = file
			elif filenames_dict[file[:file.index('_')]][file.index('_') + 1:file.index('.')] > file[file.index('_') + 1:file.index('.')]:
				filenames_dict[file[:file.index('_')]] = file

	filenames = [os.path.join(dirname, v) for k, v in filenames_dict.iteritems()]
	return filenames
	
	
def merge_xl(filenames, name, include_headers, col_start, col_end, row_start, row_end):
	
	# Create excel workbook and set active
	w = Workbook()
	ws = w.active
	
	# Initiate xl_workbook and xl_sheet objects
	xl_workbook = None
	xl_sheet = None
	
	# Initiate empty headers list
	headers = []
	
	# Offsets for pulling constrained cells (different row / col start and ends) to fit nicely into merged file
	row_offset = row_start-2 if include_headers else row_start-1
	col_offset = col_start-1
	
	# Row index, incr every file
	row_i = 0
	
	# Loop through files in filenames argument
	for file in filenames:
		# Open a workbook and set current sheet to file just opened
		print("Opening file: {}...\n".format(file))
		try:
			xl_workbook = open_workbook(file)
		except XLRDError as e:
			print e
		xl_sheet = xl_workbook.sheet_by_index(0)
		
		# Set row and column ends to either max amount or user set
		# Plus one for looping condition
		curr_row_end = xl_sheet.nrows+1 if row_end is None else row_end+1
		curr_col_end = xl_sheet.ncols+1 if col_end is None else col_end+1
		
		# Collect headers from first file, put them in the list, and write them to the file
		if include_headers:
			for c in range(col_start, curr_col_end):
				header_title = unidecode(xl_sheet.row(0)[c-1].value) if type(xl_sheet.row(0)[c-1].value) is types.UnicodeType else xl_sheet.row(0)[c-1].value
				headers.append(header_title)
				ws.cell(1, c-col_offset).value = header_title
					
			print("Got headers from file: {}\n".format(file))
			# Set to false so it doesn't rewrite every iteration
			include_headers = False
		
		# Loop through rows and columns and pull value from file and put into new file
		for r in range(row_start, curr_row_end):
			for c in range(col_start, curr_col_end):
				ws.cell(row_i + (r-row_offset), c-col_offset).value = unidecode(xl_sheet.row(r-1)[c-1].value) if type(xl_sheet.row(r-1)[c-1].value) is types.UnicodeType else xl_sheet.row(r-1)[c-1].value
				
		# Increment row index with number of rows just iterated through
		row_i += curr_row_end - row_start
	# Save the workbook to file
	w.save(name)

def move_file(name, in_dir, out_dir):
	if in_dir != out_dir:
		shutil.move(in_dir + '/' + name, out_dir + '/' + name)
	print('Saved {} to {}'.format(name, out_dir))

if __name__ == '__main__':
	try:
		opts, args = getopt.getopt(sys.argv[1:], 'x', ['name=', 'include_headers=', 'col_start=', 'col_end=', 'row_start=', 'row_end=', 'help='])
	except getopt.GetoptError as err:
		# Print help info and exit
		print(str(err)+'\n') # will print something like "option -a not recognized"
		usage()
		sys.exit(2)
	
	# If nothing passed in then print usage info
	if (len(args) == 0 and len(opts) == 0):
		usage()
		sys.exit(2)
	
	# Name of final file
	name = 'merged.xlsx'
	
	# Start column constraint
	col_start = 1
	
	# End column constraint
	col_end = None
	
	# Start row constraint
	row_start = 4
	
	# End row constraint
	row_end = None
	
	# Include headers of first file to new file
	include_headers = True
	
	for o, a in opts:
		if o == '--name':
			if any(s in '<>:"/\|?*' for s in a):
				print('{} name contains illegal characters'.format(a))
				sys.exit(2)
			elif '.' in a and a[a.index('.'):] not in ('.xls', '.xlsx'):
				print(a[a.index('.')+1:])
				print('"{}" is not of .xls or .xlsx file types'.format(a))
				sys.exit(2)
			elif '.' in a and a.index('.') > 1:
				name = a
			else:
				name = a + '.xlsx'
		if o == '--include_headers':
			include_headers = a.lower() in ('true', 't')
		if o == '--col_start':
			col_start = int(a)
		elif o == '--col_end':
			col_end = int(a)
		elif o == '--row_start':
			row_start = int(a)
		elif o == '--row_end':
			row_end = int(a)
		elif o == '--help':
			usage()
			sys.exit(2)
	
	# Get input and output directories with Tkinter
	root = Tkinter.Tk()
	root.withdraw()
	in_dir = tkFileDialog.askdirectory(title='Select Excel Input Directory')
	out_dir = tkFileDialog.askdirectory(title='Select Where to Save Merged File')
	
	if in_dir is '' or out_dir is '':
		print('Need both input and output directories\n')
		usage()
		sys.exit(2)
	
	# Get filenames
	filenames = get_filenames(in_dir)

	# Merge filenames into one excel file
	merge_xl(filenames, name, include_headers, col_start, col_end, row_start, row_end)

	# Move saved file to output directory
	move_file(name, in_dir, out_dir)